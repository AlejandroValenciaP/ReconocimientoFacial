import sqlite3
import cv2
import face_recognition as fr
import numpy as np
from datetime import datetime
import mediapipe as mp
import os
from tkinter import *
from PIL import Image, ImageTk
import imutils
import math
import openpyxl as xl

def Code_Face(images):
    listacode = []

    for img in images:
        img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        cod = fr.face_encodings(img)[0]
        listacode.append(cod)

    return listacode
def info_hora():
    inf = datetime.now()
    return inf

def hide_welcome_message():
    text.place_forget()

def regHorario(user):
    # variables
    Entrada = []
    Salida = []

    inf = info_hora()
    diasemana = inf.weekday()

    nomar = inf.strftime('%d-%m-%Y')
    texth = inf.strftime('%H:%M:%S')
    print(texth)

    print(diasemana)
    wb = xl.Workbook()

    if 12 >= inf.hour >= 6:

        if user not in Entrada:
            pos = len(Entrada)
            Entrada.append(user)
            # se guardan datos en excel
            hojam = wb.create_sheet("Entrada")
            datos = hojam.append(Entrada)
            wb.save(nomar + '.xlsx')
            si = cursor.execute(f"SELECT * FROM REGISTRO WHERE ID_USER = {user} AND FECHA = '{nomar}'")
            rows = si.fetchall()
            #print(rows)
            if (len(rows) == 0):
                cursor.execute(f"INSERT INTO REGISTRO (ID_USER,FECHA,ENTRADA) VALUES({user},'{nomar}', '{texth}')")
                conect.commit()

    else:

        if user not in Salida:
            pos = len(Salida)
            Salida.append(user)

            # se guardan datos en excel
            hojam = wb.create_sheet("Salida")
            datos = hojam.append(Salida)
            wb.save(nomar + '.xlsx')
            si = cursor.execute(f"SELECT * FROM REGISTRO WHERE ID_USER = {user} AND FECHA = '{nomar}'")
            rows = si.fetchall()
            # print(rows)
            if (len(rows) == 1):
                cursor.execute(f"UPDATE REGISTRO SET SALIDA = '{texth}' WHERE ID_USER={user} AND FECHA = '{nomar}'")
                conect.commit()


def no_coincidencia():
    global step, conteo, UserName, OutFolderPathUsers, text
    conteo = 0
    step = 0

    text = Label(screen3, text=f'No registrado')
    text.place(x=250, y=400)
    screen3.after(1000, hide_welcome_message)

def Profile():
    global step, conteo, UserName, OutFolderPathUsers, text
    conteo = 0
    step = 0

    Userfile = open(F"{OutFolderPathUsers}/{UserName}.txt")
    infoUser = Userfile.read().split(',')
    name = infoUser[0]
    user = infoUser[1]
    Pass = infoUser[2]


    if user in clases:
        text = Label(screen3, text=f'Bienvenido {name}')
        text.place(x=250, y=400)
        screen3.after(1000, hide_welcome_message)
        regHorario(user)
    print(f'yes')

def Close_Window():
    global step, conteo
    conteo = 0
    step = 0
    screen2.destroy()

def Close_Window2():
    global step, conteo
    conteo = 0
    step = 0
    screen3.destroy()


def Sign_Biometric():
    global LogUser, LogPass, OutFolderPathFaces, cap, labVideo, screen3, FaceCode, clases, images, step, parpadeos, conteo, UserName
    if cap is not None:
        ret, frame = cap.read()

        frameSave = frame.copy()

        frame = imutils.resize(frame, width=1280)

        frameRGB = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)

        frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)

        if ret == True:

            res = FaceMesh.process(frameRGB)

            px = []
            py = []
            lista = []

            if res.multi_face_landmarks:
                for rostros in res.multi_face_landmarks:
                    mpDraw.draw_landmarks(frame, rostros, FaceMeshObjet.FACE_CONNECTIONS, ConfigDraw, ConfigDraw)

                    for id, puntos in enumerate(rostros.landmark):
                        al, an, c = frame.shape
                        x, y = int(puntos.x * an), int(puntos.y * al)
                        px.append(x)
                        py.append(y)
                        lista.append([id, x, y])

                        if len(lista) == 468:
                            # ojo derecho
                            x1, y1 = lista[145][1:]
                            x2, y2 = lista[159][1:]
                            longitud1 = math.hypot(x2 - x1, y2 - y1)

                            # ojo izquierdo
                            x3, y3 = lista[374][1:]
                            x4, y4 = lista[386][1:]
                            longitud2 = math.hypot(x4 - x3, y4 - y3)

                            # parietal derecho
                            x5, y5 = lista[139][1:]
                            # parietal izq
                            x6, y6 = lista[368][1:]

                            # ceja derecha
                            x7, y7 = lista[70][1:]
                            # ceja izq
                            x8, y8 = lista[300][1:]

                            faces = detector.process(frameRGB)

                            if faces.detections is not None:
                                for face in faces.detections:
                                    score = face.score
                                    score = score[0]
                                    bbox = face.location_data.relative_bounding_box

                                    if score > cofTreshold:
                                        xi, yi, anc, alt = bbox.xmin, bbox.ymin, bbox.width, bbox.height
                                        xi, yi, anc, alt = int(xi * an), int(yi * al), int(anc * an), int(alt * al)

                                        # offsetx
                                        offsetan = (offsetx / 100) * anc
                                        xi = int(xi - int(offsetan / 2))
                                        anc = int(anc + offsetan)

                                        # offsety
                                        offsetal = (offsety / 100) * alt
                                        yi = int(yi - int(offsetal / 2))
                                        alt = int(alt + offsetal)
                                        yf = yi + alt
                                        xf = xi + anc

                                        # error
                                        if xi < 0: xi = 0
                                        if yi < 0: yi = 0
                                        if anc < 0: anc = 0
                                        if alt < 0: alt = 0

                                        # steps
                                        if step == 0:
                                            # draw
                                            cv2.rectangle(frame, (xi, yi, anc, alt), (255, 0, 255), 2)

                                            #set_image

                                            als0, ans0, c = img_step1.shape
                                            frame[50:50 + als0, 50:50 + ans0] = img_step1

                                            als1, ans1, c = img_parpadeos.shape
                                            frame[200:200 + als1, 50:50 + ans1] = img_parpadeos

                                            #face_center
                                            if x7 > x5 and x8 < x6:
                                                alch, anch, c = img_check.shape
                                                frame[100:100 + alch, 80:80 + anch] = img_check

                                                #conteo parpadeos
                                                if longitud1 <= 9 and longitud2 <= 9 and parpadeos == False:
                                                    conteo = conteo +1
                                                    parpadeos = True

                                                elif longitud1 > 10 and longitud2 > 10 and parpadeos == True:
                                                    parpadeos = False

                                                cv2.putText(frame, f'{int(conteo)}', (89,265), cv2.FONT_HERSHEY_COMPLEX, 1, (255,255,255), 1)
                                                #print(longitud1,longitud2)

                                                #conteo condicion
                                                if conteo >= 2:
                                                    alch, anch, c = img_check.shape
                                                    frame[244:244 + alch, 78:78 + anch] = img_check

                                                    #open eyes
                                                    if longitud1 > 14 and longitud2 > 14:
                                                        step = 1


                                            else:
                                                conteo = 0

                                        if step == 1:

                                            # find faces
                                            facess = fr.face_locations(frameRGB)
                                            facescod = fr.face_encodings(frameRGB, facess)

                                            for facecod, facesloc in zip(facescod, facess):
                                                # match
                                                match = fr.compare_faces(FaceCode, facecod)
                                                if match:
                                                    print(match)
                                                    # similitud
                                                    simi = fr.face_distance(FaceCode, facecod)
                                                    print(simi)
                                                    # min
                                                    min = np.argmin(simi)
                                                    if match[min]:
                                                        UserName = clases[min].upper()

                                                        Profile()
                                                    else:
                                                        print('no coincidencia')
                                                else:
                                                    no_coincidencia()
                                #close
                                close = screen3.protocol("WM_DELETE_WINDOW", Close_Window2)







                            #cv2.circle(frame, (x5, y5), 2, (255, 0, 0), cv2.FILLED)
                            #cv2.circle(frame, (x6, y6), 2, (255, 0, 0), cv2.FILLED)
                            #cv2.circle(frame, (x7, y7), 2, (255, 0, 0), cv2.FILLED)
                            #cv2.circle(frame, (x8, y8), 2, (255, 0, 0), cv2.FILLED)
                            #cv2.circle(frame, (x1, y1), 2, (255, 0, 0), cv2.FILLED)
                            #cv2.circle(frame, (x2, y2), 2, (255, 0, 0), cv2.FILLED)

        im = Image.fromarray(frame)
        img = ImageTk.PhotoImage(image=im)

        labVideo.configure(image=img)
        labVideo.image = img
        labVideo.after(10, Sign_Biometric)
    else:
        cap.release()




def log_Biometric():
    global screen2, conteo, parpadeos, img_Info, step, cap, labVideo, regUser

    if cap is not None:
        ret, frame = cap.read()

        frameSave = frame.copy()

        frame = imutils.resize(frame, width=1289)

        frameRGB = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)

        frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)

        if ret == True:

            res = FaceMesh.process(frameRGB)

            px = []
            py = []
            lista = []

            if res.multi_face_landmarks:
                for rostros in res.multi_face_landmarks:
                    mpDraw.draw_landmarks(frame, rostros, FaceMeshObjet.FACE_CONNECTIONS, ConfigDraw, ConfigDraw)

                    for id, puntos in enumerate(rostros.landmark):
                        al, an, c = frame.shape
                        x, y = int(puntos.x * an), int(puntos.y * al)
                        px.append(x)
                        py.append(y)
                        lista.append([id, x, y])

                        if len(lista) == 468:
                            # ojo derecho
                            x1, y1 = lista[145][1:]
                            x2, y2 = lista[159][1:]
                            longitud1 = math.hypot(x2 - x1, y2 - y1)

                            # ojo izquierdo
                            x3, y3 = lista[374][1:]
                            x4, y4 = lista[386][1:]
                            longitud2 = math.hypot(x4 - x3, y4 - y3)

                            # parietal derecho
                            x5, y5 = lista[139][1:]
                            # parietal izq
                            x6, y6 = lista[368][1:]

                            # ceja derecha
                            x7, y7 = lista[70][1:]
                            # ceja izq
                            x8, y8 = lista[300][1:]

                            faces = detector.process(frameRGB)

                            if faces.detections is not None:
                                for face in faces.detections:
                                    score = face.score
                                    score = score[0]
                                    bbox = face.location_data.relative_bounding_box

                                    if score > cofTreshold:
                                        xi, yi, anc, alt = bbox.xmin, bbox.ymin, bbox.width, bbox.height
                                        xi, yi, anc, alt = int(xi * an), int(yi * al), int(anc * an), int(alt * al)

                                        # offsetx
                                        offsetan = (offsetx / 100) * anc
                                        xi = int(xi - int(offsetan / 2))
                                        anc = int(anc + offsetan)

                                        # offsety
                                        offsetal = (offsety / 100) * alt
                                        yi = int(yi - int(offsetal / 2))
                                        alt = int(alt + offsetal)
                                        yf = yi + alt
                                        xf = xi + anc

                                        # error
                                        if xi < 0: xi = 0
                                        if yi < 0: yi = 0
                                        if anc < 0: anc = 0
                                        if alt < 0: alt = 0

                                        # steps
                                        if step == 0:
                                            # draw
                                            cv2.rectangle(frame, (xi, yi, anc, alt), (255, 0, 255), 2)

                                            #set_image

                                            als0, ans0, c = img_step1.shape
                                            frame[50:50 + als0, 50:50 + ans0] = img_step1

                                            als1, ans1, c = img_parpadeos.shape
                                            frame[200:200 + als1, 50:50 + ans1] = img_parpadeos

                                            #face_center
                                            if x7 > x5 and x8 < x6:
                                                alch, anch, c = img_check.shape
                                                frame[100:100 + alch, 80:80 + anch] = img_check

                                                #conteo parpadeos
                                                if longitud1 <= 9 and longitud2 <= 9 and parpadeos == False:
                                                    conteo = conteo +1
                                                    parpadeos = True

                                                elif longitud1 > 10 and longitud2 > 10 and parpadeos == True:
                                                    parpadeos = False

                                                cv2.putText(frame, f'{int(conteo)}', (89,265), cv2.FONT_HERSHEY_COMPLEX, 1, (255,255,255), 1)
                                                #print(longitud1,longitud2)

                                                #conteo condicion
                                                if conteo >= 2:
                                                    alch, anch, c = img_check.shape
                                                    frame[244:244 + alch, 78:78 + anch] = img_check


                                                    #open eyes
                                                    if longitud1 > 14 and longitud2 > 14:
                                                        print('ojos abiertos')
                                                        #cut
                                                        cut = frameSave[yi:yf, xi:xf]

                                                        #save face
                                                        cv2.imwrite(f'{OutFolderPathFaces}/{regUser}.png', cut)

                                                        step = 1



                                            else:
                                                conteo = 0

                                        if step == 1:
                                            cv2.rectangle(frame, (xi, yi, anc, alt), (0, 255, 0), 2)
                                            alli, anli, c = img_Verif.shape
                                            frame[50:50 + alli, 50:50 + anli] = img_Verif
                                #close
                                close = screen2.protocol("WM_DELETE_WINDOW", Close_Window)







                            #cv2.circle(frame, (x5, y5), 2, (255, 0, 0), cv2.FILLED)
                            #cv2.circle(frame, (x6, y6), 2, (255, 0, 0), cv2.FILLED)
                            #cv2.circle(frame, (x7, y7), 2, (255, 0, 0), cv2.FILLED)
                            #cv2.circle(frame, (x8, y8), 2, (255, 0, 0), cv2.FILLED)
                            #cv2.circle(frame, (x1, y1), 2, (255, 0, 0), cv2.FILLED)
                            #cv2.circle(frame, (x2, y2), 2, (255, 0, 0), cv2.FILLED)

        im = Image.fromarray(frame)
        img = ImageTk.PhotoImage(image=im)

        labVideo.configure(image=img)
        labVideo.image = img
        labVideo.after(10, log_Biometric)
    else:
        cap.release()


def sign():
    global  LogUser, LogPass, OutFolderPathFaces, cap, labVideo, screen3, FaceCode, clases, images

    #extract user, name
    LogUser, LogPass = inputUserReg.get(), inputPasswordReg.get()

    #DB Faces

    images = []
    clases = []
    lista = os.listdir(OutFolderPathFaces)

    #read face images
    for lis in lista:
        imgDB = cv2.imread(f'{OutFolderPathFaces}/{lis}')
        #save img
        images.append(imgDB)
        #name img
        clases.append(os.path.splitext(lis)[0])

    FaceCode = Code_Face(images)

    #new window
    screen3 = Toplevel(pantalla)
    screen3.title('Sign biometrico')
    screen3.geometry('640x480')

    labVideo = Label(screen3)
    labVideo.place(x=0, y=0)

    cap = cv2.VideoCapture(0, cv2.CAP_DSHOW)
    cap.set(3, 1280)
    cap.set(4, 720)
    Sign_Biometric()


    print('hola')


def log():
    global regName, regUser, regPassword, inputNameReg, \
        inputUserReg, inputPasswordReg, captur, labVideo, screen2, cap
    regName, regUser, regPassword = inputNameReg.get(), inputUserReg.get(), inputPasswordReg.get()

    if len(regName) == 0 or len(regUser) == 0 or len(regPassword) == 0:
        print('Formulario incompleto')
    else:
        userList = os.listdir(PathUserCheck)

        userName = []

        for lis in userList:
            user = lis
            user = user.split('.')
            userName.append((user[0]))

        if regUser in userName:
            print('Usuario registrado')

        else:
            info.append(regName)
            info.append(regUser)
            info.append(regPassword)
            cursor.execute(f"INSERT INTO USER VALUES({regUser},'{regName}', '{regPassword}')")
            conect.commit()

            f = open(f'{OutFolderPathUsers}/{regUser}.txt', "w")
            f.write(regName + ',')
            f.write(regUser + ',')
            f.write(regPassword)
            f.close()

            inputNameReg.delete(0, END)
            inputUserReg.delete(0, END)
            inputPasswordReg.delete(0, END)

            screen2 = Toplevel(pantalla)
            screen2.title('login biometrico')
            screen2.geometry('640x480')

            labVideo = Label(screen2)
            labVideo.place(x=0, y=0)

            cap = cv2.VideoCapture(0, cv2.CAP_DSHOW)
            cap.set(3, 1280)
            cap.set(4, 720)
            log_Biometric()


OutFolderPathUsers = 'Database/Users'
PathUserCheck = 'Database/Users/'
OutFolderPathFaces = 'Database/Faces'

# Read images
img_info = cv2.imread('SetUp/inicio.png')
img_step1 = cv2.imread('SetUp/paso1.png')
img_check = cv2.imread('SetUp/check.png')
img_parpadeos = cv2.imread('SetUp/parpadea.png')
img_Verif = cv2.imread('SetUp/verificacion.png')

parpadeos = False
conteo = 0
muestra = 0
step = 0

offsety = 40
offsetx = 20

cofTreshold = 0.5

mpDraw = mp.solutions.drawing_utils
ConfigDraw = mpDraw.DrawingSpec(thickness=1, circle_radius=1)

FaceMeshObjet = mp.solutions.face_mesh
FaceMesh = FaceMeshObjet.FaceMesh(max_num_faces=1)

FaceObjet = mp.solutions.face_detection
detector = FaceObjet.FaceDetection(min_detection_confidence=0.5, model_selection=1)

info = []

pantalla = Tk()
pantalla.title('RFE')
pantalla.geometry('640x720')

imageF = PhotoImage(file='SetUp/inicio.png')
background = Label(image=imageF, text='inicio')
background.place(x=0, y=0, relheight=1, relwidth=1)

inputNameReg = Entry(pantalla)
inputNameReg.place(x=75, y=313)

inputUserReg = Entry(pantalla)
inputUserReg.place(x=75, y=425)

inputPasswordReg = Entry(pantalla)
inputPasswordReg.place(x=75, y=535)
#
# inputUserdIni = Entry(pantalla)
# inputUserdIni.place(x=738,y=340)

# inputPasswordIni = Entry(pantalla)
# inputPasswordIni.place(x=738,y=460)

#Coinexion base de datos
conect = sqlite3.connect("database/bd/Reghorario")
cursor = conect.cursor()
cursor.execute("CREATE TABLE IF NOT EXISTS USER(ID_USER INTEGER PRIMARY KEY, NOMBRE VARCHAR(30),CONTRA TEXT(30))")
cursor.execute("CREATE TABLE IF NOT EXISTS REGISTRO(ID_REG INTEGER PRIMARY KEY AUTOINCREMENT, ID_USER INTEGER, FECHA TEXT(40), ENTRADA TEXT(20), SALIDA TEXT(20), FOREIGN KEY(ID_USER)REFERENCES USER(ID_USER))")


imagenLog = PhotoImage(file='SetUp/signup.png')
btReg = Button(pantalla, text='Registro', image=imagenLog, height=40, width=200, command=log)
btReg.place(x=75, y=635)

imagenSign = PhotoImage(file='SetUp/login.png')
btSign = Button(pantalla, text='Registro', image=imagenSign, height=40, width=200, command=sign)
btSign.place(x=400, y=635)

pantalla.mainloop()
